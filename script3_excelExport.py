#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
script3_excelExport.py
Export JSON -> Excel with formatting driven by ./config/script3.yaml.

Features:
- Each top-level key under "data" becomes a sheet.
- Each non-primitive sub-object becomes its own table in the parent sheet.
- Sub-objects are NOT inserted as rows in parent tables.
- Table header row can be disabled via config (show_headers: true/false).
- Option to hide Excel gridlines via config (show_gridlines: true/false).
- Optional index sheet listing all created tables (enabled by default in YAML).
- All formatting (font, sizes, colors, title fill, row fill, borders, separation rows, etc.) is read from the YAML.
"""
from pathlib import Path
import argparse
import json
from typing import Any, Dict, List, Tuple

import pandas as pd
import yaml
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# -------------------------
# Utilitaires
# -------------------------
def load_config(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}

def ensure_dir(path: str):
    Path(path).mkdir(parents=True, exist_ok=True)

def is_primitive(value: Any) -> bool:
    return isinstance(value, (str, int, float, bool)) or value is None

def normalize_table_name(path_parts: List[str]) -> str:
    return "|".join(path_parts)

def normalize_color(col: str) -> str:
    """
    Accept color strings with or without leading '#'.
    Return color string without '#' suitable for openpyxl (RRGGBB) or empty string if falsy.
    """
    if not col:
        return ""
    return col.lstrip("#")

# -------------------------
# Extraction des tableaux
# -------------------------
def extract_tables_from_obj(obj: Any, path: List[str]) -> List[Tuple[List[str], Any]]:
    tables: List[Tuple[List[str], Any]] = []

    if isinstance(obj, dict):
        primitives = {k: v for k, v in obj.items() if is_primitive(v)}
        if primitives:
            tables.append((path.copy(), {"__primitives": primitives}))
        for k, v in obj.items():
            if not is_primitive(v):
                tables.extend(extract_tables_from_obj(v, path + [k]))
    elif isinstance(obj, list):
        if len(obj) == 0:
            tables.append((path.copy(), {"__empty_list": True}))
        elif all(isinstance(x, dict) for x in obj):
            rows = []
            for idx, item in enumerate(obj):
                row = {"__index": idx}
                for k, v in item.items():
                    if is_primitive(v):
                        row[k] = v
                rows.append(row)
                for k, v in item.items():
                    if not is_primitive(v):
                        tables.extend(extract_tables_from_obj(v, path + [f"[{idx}]", k]))
            tables.append((path.copy(), {"__list_of_dicts": rows}))
        elif all(is_primitive(x) for x in obj):
            tables.append((path.copy(), {"__primitives_list": obj}))
        else:
            rows = []
            for idx, item in enumerate(obj):
                if is_primitive(item):
                    rows.append({"Index": idx, "Valeur": item})
                else:
                    tables.extend(extract_tables_from_obj(item, path + [f"[{idx}]"]))
            if rows:
                tables.append((path.copy(), {"__mixed_list_primitives": rows}))
    else:
        tables.append((path.copy(), {"__primitive_value": obj}))

    return tables

# -------------------------
# Conversion en DataFrame (lignes)
# -------------------------
def table_like_to_dataframe(table_like: Any) -> pd.DataFrame:
    if isinstance(table_like, dict):
        if "__primitives" in table_like:
            rows = [{"Attribut": k, "Valeur": v} for k, v in table_like["__primitives"].items()]
            return pd.DataFrame(rows, columns=["Attribut", "Valeur"])
        if "__list_of_dicts" in table_like:
            rows = table_like["__list_of_dicts"]
            if not rows:
                return pd.DataFrame()
            df = pd.DataFrame(rows)
            if "__index" in df.columns:
                df = df.rename(columns={"__index": "Index"})
                cols = ["Index"] + [c for c in df.columns if c != "Index"]
                df = df[cols]
            else:
                df.insert(0, "Index", range(len(df)))
            return df
        if "__primitives_list" in table_like:
            rows = [{"Index": i, "Valeur": v} for i, v in enumerate(table_like["__primitives_list"])]
            return pd.DataFrame(rows, columns=["Index", "Valeur"])
        if "__mixed_list_primitives" in table_like:
            return pd.DataFrame(table_like["__mixed_list_primitives"])
        if "__primitive_value" in table_like:
            return pd.DataFrame([{"Valeur": table_like["__primitive_value"]}], columns=["Valeur"])
        if "__empty_list" in table_like:
            return pd.DataFrame()
    return pd.DataFrame([{"Valeur": json.dumps(table_like, ensure_ascii=False)}], columns=["Valeur"])

# -------------------------
# Mise en forme centralisée
# -------------------------
def apply_table_formatting(writer, ws, start_row: int, df: pd.DataFrame, cfg: Dict[str, Any], title: str, index_entries: List[Dict[str, Any]]) -> int:
    fmt = cfg.get("format", {})
    font_name = fmt.get("font_name", "DejaVu Sans")
    font_size = fmt.get("font_size", 10)
    title_font_size = fmt.get("title_font_size", max(font_size, 11))
    title_font_color = normalize_color(fmt.get("title_font_color", "#000000"))
    title_fill = normalize_color(fmt.get("title_fill_color", "#b5bd00"))
    title_bold = fmt.get("title_bold", True)
    attributes_bold = fmt.get("attributes_bold", True)
    border_style = fmt.get("border_style", "thin")
    border_color = normalize_color(fmt.get("border_color", "#b5bd00"))
    row_fill_color = normalize_color(fmt.get("row_fill_color", "#efefef"))
    separation = int(fmt.get("table_separation_rows", 2))
    column_padding = int(fmt.get("column_padding", 2))
    show_headers = bool(fmt.get("show_headers", True))
    show_gridlines = bool(fmt.get("show_gridlines", True))

    side = Side(border_style=border_style, color=border_color if border_color else "000000")

    # Title cell
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(name=font_name, size=title_font_size, bold=title_bold, color=title_font_color)
    if title_fill:
        title_cell.fill = PatternFill(start_color=title_fill, end_color=title_fill, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Write dataframe using pandas (pandas startrow is 0-based)
    header_row = start_row + 1 if show_headers else start_row
    pandas_startrow = header_row - 1
    df.to_excel(writer, sheet_name=ws.title, index=False, startrow=pandas_startrow, header=show_headers)

    # Counts
    n_header = 1 if (show_headers and not df.empty) else 0
    n_data = len(df.index)

    # Header formatting (top border + fill + font) - ensure formatting applied after pandas wrote header
    if n_header:
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = Font(name=font_name, size=font_size, bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # top border for header
            cell.border = Border(top=side)
            # apply same row fill color as data rows if configured
            if row_fill_color:
                cell.fill = PatternFill(start_color=row_fill_color, end_color=row_fill_color, fill_type="solid")

    # Data rows formatting: bottom border only + optional row fill + attribute bolding
    data_start_row = header_row + n_header
    for r in range(n_data):
        excel_row = data_start_row + r
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            col_name = df.columns[col_idx - 1]
            bold_flag = attributes_bold and (col_name.lower() in ("attribut", "attributs", "attribute", "attribute(s)"))
            cell.font = Font(name=font_name, size=font_size, bold=bold_flag)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # bottom border for each data row (horizontal line)
            cell.border = Border(bottom=side)
            # row fill
            if row_fill_color:
                cell.fill = PatternFill(start_color=row_fill_color, end_color=row_fill_color, fill_type="solid")

    # Apply gridlines visibility for the sheet
    ws.sheet_view.showGridLines = bool(show_gridlines)

    total_rows = 1 + n_header + n_data

    # Record index entry for this table
    index_entries.append({
        "sheet": ws.title,
        "table_path": title,
        "start_row": start_row,
        "rows_count": total_rows
    })

    return total_rows

def autofit_columns(ws, cfg: Dict[str, Any]):
    padding = int(cfg.get("format", {}).get("column_padding", 2))
    min_width = int(cfg.get("format", {}).get("min_column_width", 8))
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value = cell.value
                length = 0 if value is None else len(str(value))
                if length > max_length:
                    max_length = length
            except:
                pass
        adjusted_width = max(min_width, max_length + padding)
        ws.column_dimensions[col_letter].width = adjusted_width

# -------------------------
# Ecriture Excel (orchestration)
# -------------------------
def write_tables_to_sheet_with_format(writer: pd.ExcelWriter, sheet_name: str, tables: List[Tuple[List[str], Any]], cfg: Dict[str, Any], index_entries: List[Dict[str, Any]]):
    wb = writer.book
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    writer.sheets[sheet_name] = ws

    startrow = 1  # 1-based
    separation = int(cfg.get("format", {}).get("table_separation_rows", 2))

    for path_parts, table_like in tables:
        title = normalize_table_name(path_parts)
        df = table_like_to_dataframe(table_like)
        rows_written = apply_table_formatting(writer, ws, startrow, df, cfg, title, index_entries)
        startrow += rows_written + separation

    autofit_columns(ws, cfg)

# -------------------------
# Index sheet creation
# -------------------------
def create_index_sheet(writer: pd.ExcelWriter, cfg: Dict[str, Any], index_entries: List[Dict[str, Any]]):
    if not index_entries:
        return
    wb = writer.book
    index_name = cfg.get("index_sheet_name", "Index")[: cfg.get("sheet_name_max_length", 31)]
    # If sheet exists, remove it to recreate fresh
    if index_name in wb.sheetnames:
        std = wb[index_name]
        wb.remove(std)
    # Build DataFrame
    df_index = pd.DataFrame(index_entries, columns=["sheet", "table_path", "start_row", "rows_count"])
    # Write index sheet
    df_index.to_excel(writer, sheet_name=index_name, index=False, startrow=0, header=True)
    ws = writer.book[index_name]
    # Apply simple formatting: bold header, autofit, gridlines per config
    fmt = cfg.get("format", {})
    font_name = fmt.get("font_name", "DejaVu Sans")
    font_size = fmt.get("font_size", 10)
    # Header formatting
    for col_idx, col_name in enumerate(df_index.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(name=font_name, size=font_size, bold=True)
    ws.sheet_view.showGridLines = bool(fmt.get("show_gridlines", True))
    autofit_columns(ws, cfg)

# -------------------------
# Main export
# -------------------------
def export_json_to_excel(json_path: str, output_dir: str, cfg: Dict[str, Any]) -> str:
    with open(json_path, "r", encoding="utf-8") as f:
        payload = json.load(f)

    if "data" not in payload:
        raise ValueError("Le JSON d'entrée doit contenir une clé racine 'data'.")

    data = payload["data"]
    ensure_dir(output_dir)
    output_filename = cfg.get("output_filename") or f"{Path(json_path).stem}_export.xlsx"
    out_path = Path(output_dir) / output_filename

    index_entries: List[Dict[str, Any]] = []

    with pd.ExcelWriter(out_path, engine="openpyxl") as global_writer:
        global writer
        writer = global_writer
        for top_key, top_obj in data.items():
            sheet_name = str(top_key)[: cfg.get("sheet_name_max_length", 31)]
            tables = extract_tables_from_obj(top_obj, [top_key])
            if not tables:
                df = pd.DataFrame([{"Valeur": top_obj}], columns=["Valeur"])
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.book[sheet_name]
                # apply gridlines setting for single-table sheets too
                ws.sheet_view.showGridLines = bool(cfg.get("format", {}).get("show_gridlines", True))
                autofit_columns(ws, cfg)
                # record a simple index entry for the top-level primitive
                index_entries.append({
                    "sheet": sheet_name,
                    "table_path": sheet_name,
                    "start_row": 1,
                    "rows_count": 1
                })
            else:
                write_tables_to_sheet_with_format(writer, sheet_name, tables, cfg, index_entries)

        # Create index sheet if enabled
        if cfg.get("index_sheet", True):
            create_index_sheet(writer, cfg, index_entries)

    print(f"Export terminé : {out_path}")
    return str(out_path)

# -------------------------
# CLI
# -------------------------
def main():
    parser = argparse.ArgumentParser(description="Export JSON -> Excel (format driven by YAML)")
    parser.add_argument("--config", "-c", default="./config/script3.yaml", help="Chemin du fichier de configuration YAML")
    parser.add_argument("--input", "-i", default=None, help="Chemin du fichier JSON d'entrée (optionnel, prioritaire sur le yaml)")
    parser.add_argument("--output-dir", "-o", default=None, help="Répertoire de sortie (optionnel, prioritaire sur le yaml)")
    args = parser.parse_args()

    cfg = load_config(args.config)
    json_path = args.input or cfg.get("input_path")
    output_dir = args.output_dir or cfg.get("output_dir", "./data/execlExport")

    if not json_path:
        raise SystemExit("Aucun chemin d'entrée fourni. Définissez 'input_path' dans le yaml ou utilisez --input.")

    export_json_to_excel(json_path, output_dir, cfg)

if __name__ == "__main__":
    main()
